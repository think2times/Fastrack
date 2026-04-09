from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import pandas as pd


def merge_cells_by_col(ws, col_idx):
    # 简单示例：合并第一列中内容相同的连续单元格
    start_row = 4 # 表头下第一行
    for row in range(5, ws.max_row + 1):
        if ws.cell(row, col_idx).value == ws.cell(row-1, col_idx).value:
            ws.merge_cells(start_row=row-1, start_column=col_idx, end_row=row, end_column=col_idx)

def save_styled_excel(df, config, full_path, subcoms):
    # 1. 计算合计行 (如果配置了需要计算的列)
    if 'sum_cols' in config:
        summary_row = {col: '' for col in df.columns}
        summary_row[df.columns[0]] = '合计'
        for col in config['sum_cols']:
            if col in df.columns:
                summary_row[col] = df[col].sum()
        df = pd.concat([pd.DataFrame([summary_row]), df], ignore_index=True)

    # 2. 写入 Excel
    writer = pd.ExcelWriter(full_path, engine='openpyxl')
    df.to_excel(writer, index=False, startrow=3, sheet_name='Sheet1')
    ws = writer.sheets['Sheet1']
    
    # 3. 插入标题 (合并单元格)
    # 获取报表的总列数（例如 A 到 J）
    max_col = df.shape[1]
    max_col_letter = get_column_letter(max_col)

    # 处理报表的特殊要求
    # 3-7 报表分公司列需要合并
    if config['proc'] == 'RPT_WLMQ_307':
        merge_cells_by_col(ws, 1)

    # 合并第一行单元格
    ws.merge_cells(f"A1:{max_col_letter}1")
    ws['A1'] = config['title']
    ws['A1'].font = Font(size=16, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(f"A2:{max_col_letter}2")
    ws['A2'] = f"统计账期：{datetime.now().strftime('%Y%m')}           制表时间：{datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(size=11)
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells(f"A3:{max_col_letter}3")
    # 如果 subcoms 是 None, '', ' ', 或者是全空格，都统一处理为 '全部管理站'
    display_subcoms = subcoms.strip() if (subcoms and str(subcoms).strip()) else '全部管理站'
    ws['A3'] = f"统计站点：{display_subcoms}"
    ws['A3'].font = Font(size=11)
    ws['A3'].alignment = Alignment(horizontal='center', vertical='center')

    # 4. 美化表头与数据
    header_fill = PatternFill(start_color='B0C4DE', end_color='B0C4DE', fill_type='solid')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))

    for row_idx, row in enumerate(ws.iter_rows(min_row=4, max_row=ws.max_row), start=4):
        for cell in row:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            
            # 第一行(表头)特殊处理
            if row_idx == 4:
                cell.fill = header_fill
                cell.font = Font(color='000000', bold=True)
            # 合计行加粗
            elif row_idx == 5:
                cell.font = Font(bold=True)

    # 5. 自动调整列宽
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            val = str(row[0].value)
            length = sum(2 if ord(c) > 127 else 1 for c in val)
            if length > max_len: max_len = length
        ws.column_dimensions[col_letter].width = max_len + 4

    writer.close()
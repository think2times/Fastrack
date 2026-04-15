from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# --- 1. 定义样式常量 (预加载，提高性能) ---
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
HEADER_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
CENTER_ALIGMENT = Alignment(horizontal="center", vertical="center")
LEFT_ALIGMENT = Alignment(horizontal="left", vertical="center")
RIGHT_ALIGMENT = Alignment(horizontal="right", vertical="center")

# --- 2. 纯工具函数 (无业务逻辑) ---
def merge_cells_by_col(ws, col_idx, start_row):
    """
    通用单列合并工具
    """
    current_value = ws.cell(row=start_row, column=col_idx).value
    begin_row = start_row
    max_row = ws.max_row

    for row in range(start_row + 1, max_row + 1):
        cell_value = ws.cell(row=row, column=col_idx).value
        
        # 只要不是最后一行且值相同，就继续
        if cell_value == current_value and cell_value is not None and "合计" not in str(cell_value):
            continue
        else:
            # 发现值变了，或者到了“合计”行，合并之前的区域
            if row - 1 > begin_row:
                ws.merge_cells(start_row=begin_row, start_column=col_idx, end_row=row - 1, end_column=col_idx)
                ws.cell(begin_row, col_idx).alignment = CENTER_ALIGMENT
            
            # 开启新的一组
            begin_row = row
            current_value = cell_value

    # --- 处理循环结束后，最后一组可能的遗留合并 ---
    if max_row > begin_row:
        # 检查最后这一块是否内容相同且不是合计
        last_val = ws.cell(row=begin_row, column=col_idx).value
        if last_val is not None and "合计" not in str(last_val):
            ws.merge_cells(start_row=begin_row, start_column=col_idx, end_row=max_row, end_column=col_idx)
            ws.cell(begin_row, col_idx).alignment = CENTER_ALIGMENT

def merge_nested_report_cells(ws, df, parent_col, child_col):
    """
    实现嵌套合并：只有在 parent_col(银行) 相同的区间内，才对 child_col(分公司) 进行连续合并
    """
    col_map = {col: i + 1 for i, col in enumerate(df.columns)}
    p_idx = col_map.get(parent_col)
    c_idx = col_map.get(child_col)
   
    start_row = 5 # 数据起始行
    max_row = ws.max_row

    current_p_start = start_row
    while current_p_start <= max_row:
        # --- 1. 逻辑确定父级边界 (p_end) ---
        p_val = ws.cell(row=current_p_start, column=p_idx).value
        p_end = current_p_start
        
        while p_end + 1 <= max_row:
            next_p_val = ws.cell(row=p_end + 1, column=p_idx).value
            # 关键：如果下一行是 None 或 空字符串，逻辑上认为它延续了当前的 p_val
            if next_p_val is None or str(next_p_val).strip() == "" or next_p_val == p_val:
                p_end += 1
            else:
                break
        
        # --- 2. 在确定的 [current_p_start, p_end] 区间内处理子级 ---
        c_start = current_p_start
        while c_start <= p_end:
            c_val = ws.cell(row=c_start, column=c_idx).value
            if c_val is None: # 子级如果也是 None，跳过
                c_start += 1
                continue
                
            c_run_end = c_start
            while c_run_end + 1 <= p_end:
                next_c_val = ws.cell(row=c_run_end + 1, column=c_idx).value
                # 子级同样应用逻辑填充：如果是 None 则视为与当前 c_val 相同
                if next_c_val == c_val or next_c_val is None or str(next_c_val).strip() == "":
                    c_run_end += 1
                else:
                    break
            
            if c_run_end > c_start:
                ws.merge_cells(start_row=c_start, start_column=c_idx, 
                               end_row=c_run_end, end_column=c_idx)
                ws.cell(row=c_start, column=c_idx).alignment = CENTER_ALIGMENT
            
            c_start = c_run_end + 1

        # --- 3. 最后合并父级区间 ---
        if p_end > current_p_start:
            ws.merge_cells(start_row=current_p_start, start_column=p_idx, 
                           end_row=p_end, end_column=p_idx)
            ws.cell(row=current_p_start, column=p_idx).alignment = CENTER_ALIGMENT
            
        current_p_start = p_end + 1

def apply_sheet_style(ws, df, config, sheet_name):
    """
    专门负责 Excel 穿衣服（样式）
    """
    # 0. 基础设置
    max_col = df.shape[1]
    max_col_letter = get_column_letter(max_col)
    # 如果 subcoms 是 None, '', ' ', 或者是全空格，都统一处理为 '全部分公司'
    display_subcoms = sheet_name if sheet_name != '全部' else '全部分公司'

    # 1. 标题美化
    ws.row_dimensions[1].height = 30
    ws.merge_cells(f"A1:{max_col_letter}1")
    title_text = f"{display_subcoms}{config.get('title', '')}"
    ws['A1'] = title_text
    ws['A1'].font = Font(size=20, bold=True)
    ws['A1'].alignment = CENTER_ALIGMENT

    # 2. 副标题（时间、站点）
    # 左侧：合并前一半列，设置靠左
    ws.row_dimensions[2].height = 25
    ws.merge_cells(f"A2:{get_column_letter(max_col // 2)}2")
    ws["A2"] = f"统计账期：{datetime.now().strftime('%Y%m')}"
    ws['A2'].font = Font(size=12)
    ws["A2"].alignment = LEFT_ALIGMENT

    # 右侧：合并剩余列，设置靠右
    ws.merge_cells(f"{get_column_letter(max_col // 2 + 1)}2:{max_col_letter}2")
    right_cell = ws.cell(row=2, column=max_col // 2 + 1)
    right_cell.value = f"制表时间：{datetime.now().strftime('%Y-%m-%d')}"
    right_cell.font = Font(size=12)
    right_cell.alignment = RIGHT_ALIGMENT

    ws.row_dimensions[3].height = 25
    ws.merge_cells(f"A3:{max_col_letter}3")
    ws['A3'] = f"统计站点：{display_subcoms}"
    ws['A3'].font = Font(size=12)
    ws['A3'].alignment = LEFT_ALIGMENT

    # 3. 自动合并逻辑 (只根据配置调用工具)
    merge_cols = config.get('merge_cols', [])
    if len(merge_cols) >= 2:
        merge_nested_report_cells(ws, df, merge_cols[0], merge_cols[1])
    elif merge_cols:
        # 数据从第 5 行开始（第 4 行是表头）
        merge_cells_by_col(ws, col_idx=1, start_row=5)

    # 4. 全局单元格样式与列宽计算
    is_large_data = len(df) > 10000
    
    # 预设列宽字典，减少反复调用 get_column_letter
    col_widths = {i: 0 for i in range(1, max_col + 1)}

    if is_large_data:
        # --- 大数据模式：只管表头，剩下的给默认值 ---
        print("数据量较大，采用快速格式化模式...")
        for cell in ws[4]: # 第 4 行是表头
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGMENT
            cell.fill = HEADER_FILL
            cell.font = Font(bold=True)
        
        # 统一给个默认宽度，不再逐行计算
        for i in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18
    else:
        # --- 普通模式：一次遍历完成所有事 ---
        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            curr_row = row[0].row
            ws.row_dimensions[curr_row].height = 22.5
            
            for i, cell in enumerate(row, start=1):
                # A. 设置样式
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGMENT
                if curr_row == 4:
                    cell.fill = HEADER_FILL
                    cell.font = Font(bold=True)
                
                # B. 顺便计算当前单元格内容长度，更新最大宽度
                val = str(cell.value or "")
                # 中文占2个宽度，英文占1个
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                if length > col_widths[i]:
                    col_widths[i] = length

        # 遍历结束后，统一设置列宽
        for i, width in col_widths.items():
            # 限制最大宽度，防止内容过多导致列太长
            final_width = min(max(width + 4, 10), 50) 
            ws.column_dimensions[get_column_letter(i)].width = final_width

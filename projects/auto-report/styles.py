import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from datetime import datetime
import re


def merge_cells_by_col(ws, col_idx, start_row):
    """
    确保最后一组也能正确合并
    """
    current_value = ws.cell(row=start_row, column=col_idx).value
    begin_row = start_row
    max_row = ws.max_row

    for row in range(start_row + 1, max_row + 1):
        cell_value = ws.cell(row=row, column=col_idx).value
        
        # 只要不是最后一行且值相同，就继续
        if cell_value == current_value and cell_value is not None and "合计" not in str(cell_value):
            continue
        else:
            # 发现值变了，或者到了“合计”行，合并之前的区域
            if row - 1 > begin_row:
                ws.merge_cells(start_row=begin_row, start_column=col_idx, end_row=row - 1, end_column=col_idx)
                ws.cell(begin_row, col_idx).alignment = Alignment(horizontal='center', vertical='center')
            
            # 开启新的一组
            begin_row = row
            current_value = cell_value

    # --- 处理循环结束后，最后一组可能的遗留合并 ---
    if max_row > begin_row:
        # 检查最后这一块是否内容相同且不是合计
        last_val = ws.cell(row=begin_row, column=col_idx).value
        if last_val is not None and "合计" not in str(last_val):
            ws.merge_cells(start_row=begin_row, start_column=col_idx, end_row=max_row, end_column=col_idx)
            ws.cell(begin_row, col_idx).alignment = Alignment(horizontal='center', vertical='center')

def merge_nested_report_cells(ws, df, parent_col, child_col):
    """
    实现嵌套合并：只有在 parent_col(银行) 相同的区间内，才对 child_col(分公司) 进行连续合并
    """
    col_map = {col: i + 1 for i, col in enumerate(df.columns)}
    p_idx = col_map.get(parent_col)
    c_idx = col_map.get(child_col)
   
    start_row = 5 # 数据起始行
    max_row = ws.max_row

    current_p_start = start_row
    while current_p_start <= max_row:
        # --- 1. 逻辑确定父级边界 (p_end) ---
        p_val = ws.cell(row=current_p_start, column=p_idx).value
        p_end = current_p_start
        
        while p_end + 1 <= max_row:
            next_p_val = ws.cell(row=p_end + 1, column=p_idx).value
            # 关键：如果下一行是 None 或 空字符串，逻辑上认为它延续了当前的 p_val
            if next_p_val is None or str(next_p_val).strip() == "" or next_p_val == p_val:
                p_end += 1
            else:
                break
        
        # --- 2. 在确定的 [current_p_start, p_end] 区间内处理子级 ---
        c_start = current_p_start
        while c_start <= p_end:
            c_val = ws.cell(row=c_start, column=c_idx).value
            if c_val is None: # 子级如果也是 None，跳过
                c_start += 1
                continue
                
            c_run_end = c_start
            while c_run_end + 1 <= p_end:
                next_c_val = ws.cell(row=c_run_end + 1, column=c_idx).value
                # 子级同样应用逻辑填充：如果是 None 则视为与当前 c_val 相同
                if next_c_val == c_val or next_c_val is None or str(next_c_val).strip() == "":
                    c_run_end += 1
                else:
                    break
            
            if c_run_end > c_start:
                ws.merge_cells(start_row=c_start, start_column=c_idx, 
                               end_row=c_run_end, end_column=c_idx)
                ws.cell(row=c_start, column=c_idx).alignment = Alignment(vertical='center', horizontal='center')
            
            c_start = c_run_end + 1

        # --- 3. 最后合并父级区间 ---
        if p_end > current_p_start:
            ws.merge_cells(start_row=current_p_start, start_column=p_idx, 
                           end_row=p_end, end_column=p_idx)
            ws.cell(row=current_p_start, column=p_idx).alignment = Alignment(vertical='center', horizontal='center')
            
        current_p_start = p_end + 1

def apply_sheet_style(ws, df, config, sheet_name):
    # 0. 基础设置
    max_col = df.shape[1]
    max_col_letter = get_column_letter(max_col)
    # 如果 subcoms 是 None, '', ' ', 或者是全空格，都统一处理为 '全部管理站'
    display_subcoms = sheet_name if sheet_name != '全部' else '全部分公司'

    # 1. 标题美化
    ws.merge_cells(f"A1:{max_col_letter}1")
    ws['A1'] = display_subcoms + config.get('title', '') if display_subcoms != '全部' else '' + config['title']
    ws['A1'].font = Font(size=20, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 30

    # 左侧：合并前一半列，设置靠左
    ws.merge_cells(f"A2:{get_column_letter(max_col // 2)}2")
    ws["A2"] = f"统计账期：{datetime.now().strftime('%Y%m')}"
    ws['A2'].font = Font(size=12)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")

    # 右侧：合并剩余列，设置靠右
    ws.merge_cells(f"{get_column_letter(max_col // 2 + 1)}2:{max_col_letter}2")
    right_cell = ws.cell(row=2, column=max_col // 2 + 1)
    right_cell.value = f"制表时间：{datetime.now().strftime('%Y-%m-%d')}"
    right_cell.font = Font(size=12)
    right_cell.alignment = Alignment(horizontal="right", vertical="center")
    ws.row_dimensions[2].height = 25

    ws.merge_cells(f"A3:{max_col_letter}3")
    ws['A3'] = f"统计站点：{display_subcoms}"
    ws['A3'].font = Font(size=12)
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[3].height = 25

    # 2. 处理带有小计的报表的合并逻辑
    merge_cols = config.get('merge_cols', '')
    if merge_cols and len(merge_cols) > 1:
        # merge_report_cells(ws, df, merge_cols)
        merge_nested_report_cells(ws, df, merge_cols[0], merge_cols[1])
    elif merge_cols:
        # 数据从第 5 行开始（第 4 行是表头）
        merge_cells_by_col(ws, col_idx=1, start_row=5)

    # 3. 通用格式：边框、居中、表头颜色
    if len(df) > 10000:
        # 大数据模式：仅处理表头和合计行（即前两行数据），不循环全表
        print("数据量较大，采用快速格式化模式...")
        
        # 只美化表头（第4行）
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        for cell in ws[4]:
            cell.fill = header_fill
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 设置一个默认的全局行高和列宽，不再动态计算
        ws.row_dimensions[4].height = 22.5
        for col_idx in range(1, max_col + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
    else:
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), 
                             top=Side(style="thin"), bottom=Side(style="thin"))
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")

        for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
            # 获取当前行号
            row_num = row[0].row
    
            # 设置该行行高
            ws.row_dimensions[row_num].height = 22.5

            # 设置该行所有单元格的样式
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                # 表头样式
                if cell.row == 4:
                    cell.fill = header_fill
                    cell.font = Font(bold=True)

        # 自动调整列宽
        for col_idx in range(1, max_col + 1):
            col_letter = get_column_letter(col_idx)
            max_len = 0
            for row in ws.iter_rows(min_row=4, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
                val = str(row[0].value)
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                if length > max_len: 
                    max_len = length
            ws.column_dimensions[col_letter].width = max_len + 4

def clean_illegal_chars(val):
    if isinstance(val, str):
        # 匹配所有 Excel 不支持的控制字符
        # \x00-\x08, \x0b-\x0c, \x0e-\x1f
        ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')
        return ILLEGAL_CHARACTERS_RE.sub("", val)
    return val

def export_sheet_excel(df, config, full_path):
    """
    专门为 3-11 这种需要按分公司拆分 Sheet 的报表设计的导出函数
    """

    # 使用 with 语句，确保即使报错也会保存文件，防止“空表”
    with pd.ExcelWriter(full_path, engine="openpyxl") as writer:
        # --- 第 1 个 Sheet: 全部数据 ---
        # 先计算全表的合计行（逻辑复用你之前的）
        all_df = generate_summary_df(df, config) 

        # 导出前执行清洗
        all_df = all_df.map(clean_illegal_chars)
        all_df.to_excel(writer, index=False, startrow=3, sheet_name="全部")
        apply_sheet_style(writer.sheets["全部"], all_df, config, "全部")

        # --- 第 2-N 个 Sheet: 按分公司拆分 ---
        # 在导出逻辑中，先根据映射表找到对应的中文列名
        real_split_col = config.get('split_by')
        # 如果 columns_map 里有这个 key，就取它的 value（即“分公司”）
        display_split_col = config['columns_map'].get(real_split_col, real_split_col)
        if display_split_col and display_split_col in df.columns:
            groups = df.groupby(display_split_col)
            for name, group_df in groups:
                # 过滤掉无效或空的名字
                sheet_name = str(name)[:12]     # 最长为12个字符

                # 为当前分公司计算单独的合计行
                current_df = generate_summary_df(group_df, config)
                current_df.to_excel(writer, index=False, startrow=3, sheet_name=sheet_name)

                # 应用样式
                apply_sheet_style(writer.sheets[sheet_name], current_df, config, sheet_name)

def generate_summary_df(df, cfg, sum_cols=None):
    """
    report_df: 处理完后的 DataFrame
    cfg: 当前报表的字典配置
    sum_cols: 需要求和的列名列表
    """

    if cfg.get('proc') == 'RPT_WLMQ_306':
        return df

    # 准备数据
    report_df = df.copy()

    # 如果没传 sum_cols，就自动从 cfg 里拿
    if sum_cols is None:
        sum_cols = cfg.get('sum_cols', [])
    
    if 'sum_cols' in cfg:
        summary_row = {}
    
        # 动态获取分组列（用于排除小计）
        group_col_raw = cfg.get('group_by')
        # 匹配 columns_map 中的显示名
        dim_col = cfg['columns_map'].get(group_col_raw, report_df.columns[0])
    
        # 排除所有包含“小计”的行
        if dim_col not in report_df.columns:
            # 如果找不到列，就认为每一行都是明细，不做排除
            mask = [True] * len(report_df)
        else:
            mask = ~report_df[dim_col].astype(str).str.contains('小计', na=False)

        for col in sum_cols:
            if col in report_df.columns:
                # 只有在明细行中进行 sum
                summary_row[col] = report_df.loc[mask, col].sum()

        summary_row[dim_col] = '合计'

        # 特殊逻辑：合计行单价重算 (针对 3-11)
        # 使用的是 columns_map 映射后的中文列名：'应收费用' 和 '水量'
        if cfg.get('proc') == 'RPT_WLMQ_311':
            total_money = summary_row.get('六费合计', 0)
            total_water = summary_row.get('水量', 0)
            # 防止除以 0
            summary_row['单价'] = round(total_money / total_water, 2) if total_water != 0 else 0

        # 使用 report_df 的列名构造，并同步数据类型以消除警告
        summary_df = pd.DataFrame([summary_row], columns=report_df.columns)

        # 准备待合并列表
        # 根据是否配置 sum_first 决定合计行放在最前（sum_first）还是最后
        to_concat = [summary_df, report_df] if cfg.get('sum_first') else [report_df, summary_df]

        # 核心过滤逻辑：只保留非空的 DataFrame
        valid_dfs = [d for d in to_concat if d is not None and not d.empty]

        if valid_dfs:
            report_df = pd.concat(valid_dfs, ignore_index=True)
            pass

    return report_df

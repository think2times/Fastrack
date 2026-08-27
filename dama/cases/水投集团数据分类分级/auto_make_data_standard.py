import json
import os
import re
import time
from openai import OpenAI
from docx import Document
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ==============================================================================
# 1. 基础配置
# ==============================================================================

# 数据库连接配置 (Oracle)
DB_CONFIG = {
    "user": "CSM",  # 数据库账号
    "password": "ghi22OGJELOB2f",  # 数据库密码
    "dsn": "172.16.16.11:1521/CSM_BM",  # 数据库连接地址与服务名
}

# 文件与路径配置
WORD_FILE_PATH = r"F:\Projects\Fastrack\dama\cases\水投集团数据分类分级\1-营收系统\营收系统数据分类分级标准.docx"  # 输入 Word 文档路径
BASE_OUTPUT_DIR = r"F:\Projects\Fastrack\dama\cases\营收系统数据标准\服务类"  # 主输出根目录

# ==============================================================================
# 2. AI 客户端初始化与 Prompt 定义
# ==============================================================================

# 本地部署的 Qwen3.6-27B OpenAI 兼容 API 配置
# ==================== 1. 应用平台 API 配置 ====================
APP_CODE = "e8b9ed13-9547-47b1-8cd2-6b8ab6783513"  # 替换为你的真实 app_code

ai_client = OpenAI(
    base_url="http://172.16.7.1:30080/v1",
    api_key=APP_CODE,  # 将应用代码作为 API Key
    max_retries=1,# 内部重试次数
    timeout=120.0,  # 🔥 将客户端超时时间延长至 120 秒，防止网关因等待过久断开
)

AI_MODEL_NAME = "Qwen3.6-27B"  # 替换为实际模型名称

# 针对数据治理和数据安全分类分级的系统级 Prompt
SYSTEM_PROMPT = """
你是一位水务行业资深的数据治理与数据标准化专家，擅长数据元数据梳理、数据安全分类分级。
请根据用户提供的【表信息】和【字段列表】，按照以下规范生成每个字段的标准元数据属性。

### 输出数据规范：
1. **businessMeaning (业务含义)**：用专业、精准的中文说明字段代表的业务概念（如：“标识水价变更的生效起始日期”）。
2. **businessRule (业务规则)**：说明字段在业务上的校验规则、约束条件或枚举取值逻辑（如：“采用 YYYY-MM-DD 格式，且不得早于当前日期”）。
3. **standardRule (字符规范)**：说明字符输入格式限制（如：“11位数字”、“支持中文/自由文本”、“ISO-8601日期”）。
4. **rangeRule (值域范围)**：值域取值空间（如：“正整数”、“0-未生效, 1-已生效, 2-已废止”、“自由文本”）。
5. **sensitiveLevel (敏感级别)**：严格按照以下标准判定，仅允许输出 ["低", "中", "高"]：
   - 高：涉及个人敏感隐私（身份证、手机号、银行卡、密码、精准住址等）或核心商业机密。
   - 中：涉及个人基本信息（姓名、邮箱）、企业经营核心数据、客户用水量/账单金额。
   - 低：系统自增ID、公共字典编码、配置参数、业务状态标识、时间戳、操作日志。
6. **sensitiveLabel (敏感标签)**：仅允许输出以下之一或组合 ["个人敏感信息", "个人基本信息", "一般业务数据", "国家机密信息"]。
7. **dataUnit (计量单位)**判定规则：
    - 与金额/资金/费用相关的字段（如：应收金额、水费、滞纳金、保证金等），计量单位优先填写："元"
    - 与水量/体积相关的字段（如：水量、用量、抄表水量等），计量单位优先填写："立方米"
    - 与单价/水价/费率相关的字段（如：基本水价、终端水价、单价、污水处理费率等），计量单位优先填写："元/立方米"（即每吨水的价格）
    - 编号、名称、时间、状态、类型等无物理单位的字段，计量单位统一保持为空字符串：""

### 输出格式要求：
必须返回合法的 JSON 对象，Key 为字段英文名 (COLUMN_NAME)，Value 为对应的属性对象。格式如下：
{
  "COLUMN_NAME_1": {
    "businessMeaning": "...",
    "businessRule": "...",
    "standardRule": "...",
    "rangeRule": "...",
    "sensitiveLevel": "高/中/低",
    "sensitiveLabel": "..."
  }
}
"""


def sanitize_filename(name):
    """去除文件名和文件夹名中的非法字符，防止路径创建失败"""
    return re.sub(r'[\\/*?:"<>|]', "_", str(name)).strip()


# ==============================================================================
# 3. AI 智能增强模块 (兼容 Qwen3.6-27B 思考标签清洗)
# ==============================================================================
def enhance_table_with_ai(
    level1, level2, level3, table_name, table_cn_name, columns_list, max_retries=3
):
    """调用本地 Qwen3.6-27B 一次性推导整张表所有字段的治理属性"""
    cols_summary = [
        {
            "COLUMN_NAME": c["COLUMN_NAME"],
            "DATA_TYPE": c["TYPE_FMT"],
            "NULLABLE": c["NULLABLE"],
            "COMMENTS": c["COLUMN_COMMENTS"] or "",
        }
        for c in columns_list
    ]

    user_prompt = f"""
请分析以下数据库表结构并生成数据标准属性：
- 一级分类：{level1}
- 二级分类：{level2}
- 三级分类：{level3}
- 数据库表名：{table_name}
- 表中文名称：{table_cn_name}

字段列表：
{json.dumps(cols_summary, ensure_ascii=False, indent=2)}
"""

    # 尝试重试机制
    for attempt in range(1, max_retries + 1):
        try:
            # 🔥 请求前休眠 1 秒，给本地模型释放显存/缓解排队压力
            time.sleep(1)

            response = ai_client.chat.completions.create(
                model=AI_MODEL_NAME,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": user_prompt},
                ],
                timeout=120.0,  # 显式单次请求超时
            )

            msg = response.choices[0].message

            if hasattr(msg, "reasoning_content") and msg.reasoning_content:
                print(
                    f"    🧠 [{table_name} 思考逻辑]: {msg.reasoning_content[:60]}..."
                )

            result_text = msg.content.strip() if msg.content else ""

            # 清洗 Markdown 包裹标记
            result_text = re.sub(
                r"^```json\s*", "", result_text, flags=re.IGNORECASE
            )
            result_text = re.sub(r"^```\s*", "", result_text)
            result_text = re.sub(r"\s*```$", "", result_text)

            ai_data = json.loads(result_text.strip())
            print(
                f"    🤖 [AI 增强成功] 表 {table_name} 完成元数据生成。"
            )
            return ai_data

        except Exception as e:
            if "502" in str(e) and attempt < max_retries:
                print(
                    f"    ⚠️ [502 报错] 表 {table_name} 触发网关超时/过载，正在尝试第 {attempt}/{max_retries} 次重试..."
                )
                time.sleep(3)  # 发生 502 时，等待 3 秒再重试
            else:
                print(
                    f"    ⚠️ [AI 增强降级] 表 {table_name} 重试失败 ({e})，降级使用基础规则。"
                )
                return None

    return None


# ==============================================================================
# 4. Word 文档结构解析
# ==============================================================================
def parse_word_schema(doc_path):
    """只抓取和解析‘服务类’下的二级分类、三级分类及对应数据库表"""
    doc = Document(doc_path)

    schema = {}
    curr_l1 = ""
    curr_l2 = ""
    curr_l3 = ""

    # 正则表达式定义
    # 匹配任意一级标题（如：1. 基础类、2. 服务类、3. 生产类、4. 管理类 等）
    l1_any_re = re.compile(
        r"^([1-9一二三四五六七八九十]+[\s\.、]*|\s*)([^\s]+类)"
    )
    # 匹配二级标题（如：2.1 基础档案数据、3.1 生产运行数据）
    l2_re = re.compile(r"^[1-9][0-9]*[\.\．](\d+)[\s\.]*(.+)")
    # 匹配三级标题（如：2.1.1 水价信息、3.1.1 压力数据）
    l3_re = re.compile(r"^[1-9][0-9]*[\.\．](\d+)[\.\．](\d+)[\s\.]*(.+)")

    for element in doc.element.body:
        if isinstance(element, CT_P):
            p = Paragraph(element, doc)
            text = p.text.strip()
            style_name = p.style.name.lower() if p.style else ""

            if not text:
                continue

            # 1. 判定是否到达了任意“一级标题”
            is_l1_heading = (
                ("1" in style_name or "heading 1" in style_name or "标题 1" in style_name)
                and ("类" in text or l1_any_re.match(text))
            ) or bool(l1_any_re.match(text) and not l2_re.match(text))

            if is_l1_heading:
                if "服务类" in text:
                    curr_l1 = "服务类"  # 开启抓取开关
                else:
                    # 🔥【关键修正】：一旦遇到 生产类、管理类 或其他一级标题，立即清空并切断抓取
                    curr_l1 = "OTHER"
                    curr_l2 = ""
                    curr_l3 = ""
                continue

            # 2. 只有当前处于“服务类”作用域内，才继续解析二三级标题
            if curr_l1 == "服务类":
                m3 = l3_re.match(text)
                is_l3_style = (
                    "3" in style_name
                    or "heading 3" in style_name
                    or "标题 3" in style_name
                )

                m2 = l2_re.match(text)
                is_l2_style = (
                    "2" in style_name
                    or "heading 2" in style_name
                    or "标题 2" in style_name
                )

                if m3 or (is_l3_style and not is_l2_style):
                    curr_l3 = text
                    if curr_l2 and curr_l2 not in schema:
                        schema[curr_l2] = {}
                    if curr_l2 and curr_l3 not in schema[curr_l2]:
                        schema[curr_l2][curr_l3] = []

                elif m2 or is_l2_style:
                    curr_l2 = text
                    curr_l3 = ""
                    if curr_l2 not in schema:
                        schema[curr_l2] = {}

        elif isinstance(element, CT_Tbl):
            # 3. 只有当处于“服务类”且定位到具体的二级、三级分类时，才收集表格里的数据库表
            if curr_l1 == "服务类" and curr_l2 and curr_l3:
                tbl = Table(element, doc)
                for row in tbl.rows[1:]:  # 跳过表头
                    cells = [c.text.strip() for c in row.cells]
                    if len(cells) >= 3:
                        tbl_name = cells[1].upper()
                        tbl_cn = cells[2]
                        if tbl_name and tbl_name not in ["表", "表名", "TABLE"]:
                            schema[curr_l2][curr_l3].append(
                                {
                                    "table_name": tbl_name,
                                    "table_cn_name": tbl_cn,
                                }
                            )

    return schema


# ==============================================================================
# 5. Oracle 元数据查询与字段推导
# ==============================================================================
def fetch_table_columns_from_db(cursor, table_name):
    """执行 Oracle SQL 查询表字段及其注释"""
    sql = """
    SELECT 
        t.table_name                                    AS 表名,
        t.COMMENTS                                      AS 中文表名,
        c.COLUMN_NAME                                   AS COLUMN_NAME,
        CASE 
            WHEN c.DATA_TYPE IN ('VARCHAR2', 'CHAR', 'NVARCHAR2', 'NCHAR') THEN c.DATA_TYPE || '(' || c.DATA_LENGTH || ')'
            WHEN c.DATA_TYPE = 'NUMBER' AND c.DATA_PRECISION IS NOT NULL AND c.DATA_SCALE IS NOT NULL AND c.DATA_SCALE > 0 
                THEN c.DATA_TYPE || '(' || c.DATA_PRECISION || ',' || c.DATA_SCALE || ')'
            WHEN c.DATA_TYPE = 'NUMBER' AND c.DATA_PRECISION IS NOT NULL 
                THEN c.DATA_TYPE || '(' || c.DATA_PRECISION || ')'
            ELSE c.DATA_TYPE 
        END                                             AS 数据类型格式,
        c.DATA_TYPE                                     AS DATA_TYPE,
        c.DATA_LENGTH                                   AS DATA_LENGTH,
        c.DATA_PRECISION                                AS DATA_PRECISION,
        c.DATA_SCALE                                    AS DATA_SCALE,
        c.NULLABLE                                      AS NULLABLE,
        m.COMMENTS                                      AS COLUMN_COMMENTS
    FROM 
        USER_TAB_COLUMNS c
    LEFT JOIN 
        USER_COL_COMMENTS m 
        ON c.TABLE_NAME = m.TABLE_NAME 
       AND c.COLUMN_NAME = m.COLUMN_NAME
    LEFT JOIN 
        USER_TAB_COMMENTS t
        ON c.TABLE_NAME = t.TABLE_NAME
    WHERE 
        c.TABLE_NAME = UPPER(:tbl_name)
    ORDER BY 
        c.COLUMN_ID
    """
    cursor.execute(sql, tbl_name=table_name)
    rows = cursor.fetchall()

    return [
        {
            "TABLE_NAME": r[0],
            "TABLE_COMMENTS": r[1],
            "COLUMN_NAME": r[2],
            "TYPE_FMT": r[3],
            "DATA_TYPE": r[4],
            "DATA_LENGTH": r[5],
            "DATA_PRECISION": r[6],
            "DATA_SCALE": r[7],
            "NULLABLE": r[8],
            "COLUMN_COMMENTS": r[9],
        }
        for r in rows
    ]


def derive_attributes_smart(
    row_data, level2, level3, tbl_cn_name, ai_table_cache=None
):
    """结合 AI 推导与业务规则组装 29 项数据标准元数据"""
    col_name = str(row_data["COLUMN_NAME"]).upper()
    base_type = str(row_data["DATA_TYPE"] or "").upper()
    length = row_data["DATA_LENGTH"]
    prec = row_data["DATA_PRECISION"]
    scale = row_data["DATA_SCALE"]
    nullable = row_data["NULLABLE"]
    comment = row_data["COLUMN_COMMENTS"] or ""

    # 清除分类标题开头的数字，如 "2.1 基础档案数据" -> "基础档案数据"
    l2_pure = re.sub(r"^2\.\d+\s*", "", level2).strip()
    l3_pure = re.sub(r"^2\.\d+\.\d+\s*", "", level3).strip()

    # 规则：如果数据格式为文本，则数据类型统一为 VARCHAR2
    if any(
        t in base_type for t in ["VARCHAR", "CHAR", "NVARCHAR", "NCHAR", "TEXT"]
    ):
        final_data_type = "VARCHAR2"
        data_fmt = "文本"
    elif "NUMBER" in base_type or "INT" in base_type or "DECIMAL" in base_type:
        final_data_type = "NUMBER"
        data_fmt = "数值"
    elif "DATE" in base_type or "TIME" in base_type:
        final_data_type = base_type
        data_fmt = "日期时间"
    else:
        final_data_type = "VARCHAR2"
        data_fmt = "文本"

    # 读取 AI 结果，若无 AI 则降级使用基本逻辑
    ai_col = (
        ai_table_cache.get(col_name, {})
        if (ai_table_cache and isinstance(ai_table_cache, dict))
        else {}
    )

    meaning = ai_col.get(
        "businessMeaning",
        f"标识{tbl_cn_name}的{comment if comment else col_name}",
    )
    rule = ai_col.get("businessRule", "按照实际业务数据填写")
    std_rule = ai_col.get("standardRule", "支持中文、自由文本")
    range_rule = ai_col.get("rangeRule", comment if comment else "自由文本")
    sensitive_level = ai_col.get("sensitiveLevel", "低")
    sensitive_label = ai_col.get("sensitiveLabel", "系统基础数据")

    data_len = prec if prec is not None else (length if length else "")

    return {
        "code": f"STD_{col_name}",
        "cname": comment.split("（")[0].split("(")[0] if comment else col_name,
        "ename": col_name,
        "subject": "营收系统数据标准",
        "level1": "服务类",  # 一级分类固定为服务类
        "level2": l2_pure,  # 二级分类
        "level3": l3_pure,  # 三级分类
        "businessMeaning": meaning,
        "businessRule": rule,
        "referenceDoc": "",
        "referenceCode": "",
        "subordinateDepartment": "供排水经营部",
        "referenceNo": "",
        "authoritySystem": "营业收费系统",
        "cooperativeDepartment": "信息和网络管理中心",
        "dataType": final_data_type,
        "dataFormat": data_fmt,
        "dataLength": data_len,
        "dataScale": scale if scale is not None else "",
        "dataUnit": "",
        "archiveWay": "数据库在线归档",
        "archiveScheme": "满足归档条件自动归档",
        "archivePattern": "不可销毁",
        "destroyWay": "不销毁",
        "destroyScheme": "不执行销毁，仅通过业务状态标记停用",
        "standardRule": std_rule,
        "rangeRule": range_rule,
        "nullRule": "否" if nullable == "N" else "是",
        "sensitiveLevel": sensitive_level,
        "sensitiveLabel": sensitive_label,
    }


# ==============================================================================
# 6. Excel 导出引擎 (含样式美化与超链接)
# ==============================================================================
def generate_excel_for_l3(level2, level3, table_list, cursor, l2_folder_path):
    """为三级小类生成 Excel 文件，每张表作为一个 Sheet"""
    clean_l3_name = sanitize_filename(level3)
    file_path = os.path.join(
        l2_folder_path, f"{clean_l3_name}_数据标准数据字典.xlsx"
    )

    wb = openpyxl.Workbook()

    # 1. 创建【数据字典目录】Sheet
    ws_index = wb.active
    ws_index.title = "数据字典目录"
    ws_index.views.sheetView[0].showGridLines = True

    # 顶部大标题
    ws_index.merge_cells("A1:D1")
    t_cell = ws_index["A1"]
    t_cell.value = f"【{level2} - {level3}】数据标准清单"
    t_cell.font = Font(
        name="Microsoft YaHei", size=13, bold=True, color="FFFFFF"
    )
    t_cell.fill = PatternFill(
        start_color="1F4E78", end_color="1F4E78", fill_type="solid"
    )
    t_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws_index.row_dimensions[1].height = 32

    ws_index.append(["序号", "数据库表名", "中文表名", "字段数量"])
    ws_index.row_dimensions[2].height = 22

    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for col_i in range(1, 5):
        c = ws_index.cell(row=2, column=col_i)
        c.font = Font(name="Microsoft YaHei", size=10, bold=True)
        c.fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    headers = [
        ("业务属性", "标准编号", "code", "C6EFCE"),
        ("业务属性", "标准中文标题", "cname", "C6EFCE"),
        ("业务属性", "标准英文标题", "ename", "C6EFCE"),
        ("业务属性", "标准主题", "subject", "C6EFCE"),
        ("业务属性", "一级分类", "level1", "C6EFCE"),
        ("业务属性", "二级分类", "level2", "C6EFCE"),
        ("业务属性", "三级分类", "level3", "C6EFCE"),
        ("业务属性", "业务含义", "businessMeaning", "C6EFCE"),
        ("业务属性", "业务规则", "businessRule", "C6EFCE"),
        ("管理属性", "参考文档", "referenceDoc", "BDD7EE"),
        ("管理属性", "引用部门", "referenceCode", "BDD7EE"),
        ("管理属性", "引用编号", "coordinateDepartment", "BDD7EE"),
        ("管理属性", "权威系统", "referenceAuthoritySystem", "BDD7EE"),
        ("管理属性", "协办部门", "cooperativeDepartment", "BDD7EE"),
        ("技术属性", "数据类型", "dataType", "F8CBAD"),
        ("技术属性", "数据格式", "dataFormat", "F8CBAD"),
        ("技术属性", "数据长度", "dataLength", "F8CBAD"),
        ("技术属性", "小数位数", "dataScale", "F8CBAD"),
        ("技术属性", "计量单位", "dataUnit", "F8CBAD"),
        ("生命周期属性", "归档方式", "archiveWay", "A9D08E"),
        ("生命周期属性", "归档方案", "archiveScheme", "A9D08E"),
        ("生命周期属性", "归档模式", "archivePattern", "A9D08E"),
        ("生命周期属性", "销毁方式", "destroyWay", "A9D08E"),
        ("生命周期属性", "销毁方案", "destroyScheme", "A9D08E"),
        ("质量属性", "字符规范", "standardRule", "9BC2E6"),
        ("质量属性", "值域范围", "rangeRule", "9BC2E6"),
        ("质量属性", "不允许为空", "nullRule", "9BC2E6"),
        ("安全属性", "敏感级别", "sensitiveLevel", "F4B084"),
        ("安全属性", "敏感标签", "sensitiveLabel", "F4B084"),
    ]

    top_cat = [h[0] for h in headers]
    cn_titles = [h[1] for h in headers]
    en_codes = [h[2] for h in headers]

    sheet_count = 0
    for item in table_list:
        tbl_name = item["table_name"]
        columns_data = fetch_table_columns_from_db(cursor, tbl_name)
        if not columns_data:
            print(f"    ⚠️ [数据库无此表] {tbl_name}，跳过导出。")
            continue

        sheet_count += 1
        tbl_cn_name = columns_data[0]["TABLE_COMMENTS"] or item["table_cn_name"]
        sheet_title = tbl_name[:31]  # Sheet 名字限制 31 字符

        # 写入目录行
        curr_row = sheet_count + 2
        ws_index.append([sheet_count, tbl_name, tbl_cn_name, len(columns_data)])

        ws_index.cell(row=curr_row, column=1).alignment = Alignment(
            horizontal="center"
        )
        ws_index.cell(row=curr_row, column=2).font = Font(
            name="Microsoft YaHei", color="0000FF", underline="single"
        )
        ws_index.cell(
            row=curr_row, column=2
        ).hyperlink = f"#'{sheet_title}'!A1"  # 目录超链接跳转
        ws_index.cell(row=curr_row, column=3).alignment = Alignment(
            horizontal="left"
        )
        ws_index.cell(row=curr_row, column=4).alignment = Alignment(
            horizontal="center"
        )

        for c_i in range(1, 5):
            ws_index.cell(row=curr_row, column=c_i).border = thin_border

        # 调用 AI 进行整张表的语义与敏感级别推导
        ai_table_cache = enhance_table_with_ai(
            level1="服务类",
            level2=level2,
            level3=level3,
            table_name=tbl_name,
            table_cn_name=tbl_cn_name,
            columns_list=columns_data,
        )

        # 2. 为该表创建独立的 Sheet
        ws = wb.create_sheet(title=sheet_title)
        ws.views.sheetView[0].showGridLines = True
        ws.freeze_panes = "C4"  # 冻结前3行和前2列

        ws.append(top_cat)
        ws.append(cn_titles)
        ws.append(en_codes)

        # 合并第一行属性大类
        start_col = 1
        for i in range(1, len(top_cat)):
            if top_cat[i] != top_cat[i - 1]:
                ws.merge_cells(
                    start_row=1,
                    start_column=start_col,
                    end_row=1,
                    end_column=i,
                )
                start_col = i + 1
        ws.merge_cells(
            start_row=1,
            start_column=start_col,
            end_row=1,
            end_column=len(top_cat),
        )

        # 填充字段数据
        for col_row in columns_data:
            proc = derive_attributes_smart(
                col_row, level2, level3, tbl_cn_name, ai_table_cache
            )
            ws.append([proc[h[2]] for h in headers])

        # 设置表头样式与色彩
        for col_idx, h in enumerate(headers, 1):
            fill = PatternFill(
                start_color=h[3], end_color=h[3], fill_type="solid"
            )
            for r_idx in range(1, 4):
                c = ws.cell(row=r_idx, column=col_idx)
                c.fill = fill
                c.alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
                c.font = Font(name="Microsoft YaHei", size=9, bold=(r_idx < 3))
                c.border = thin_border

        # 设置数据行样式
        for row in ws.iter_rows(
            min_row=4,
            max_row=len(columns_data) + 3,
            min_col=1,
            max_col=len(headers),
        ):
            for c in row:
                c.alignment = Alignment(
                    horizontal="left", vertical="center", wrap_text=True
                )
                c.font = Font(name="Microsoft YaHei", size=9)
                c.border = thin_border

        # 自动调整列宽
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len * 1.3, 12)

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[3].height = 18

    # 调整目录页列宽
    for col in ws_index.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_index.column_dimensions[col_letter].width = max(max_len * 1.5, 14)

    if sheet_count > 0:
        wb.save(file_path)
        print(f"  📄 [Excel 已生成] {file_path} (含 {sheet_count} 张表)")


# ==============================================================================
# 7. 主流程控制入口
# ==============================================================================
import oracledb


def main():
    if not os.path.exists(BASE_OUTPUT_DIR):
        os.makedirs(BASE_OUTPUT_DIR)

    print("1. 正在解析 Word 架构树...")
    schema = parse_word_schema(WORD_FILE_PATH)
    print(f"   解析完成！识别到 {len(schema)} 个二级分类。\n")

    print("2. 正在初始化 Oracle 客户端 (Thick 模式)...")
    try:
        # 修改为 Instant Client 实际解压路径：
        INSTANT_CLIENT_DIR = r"F:\app\pluto\product\instantclient_19_32"

        oracledb.init_oracle_client(lib_dir=INSTANT_CLIENT_DIR)
        print("   ✅ 19c Oracle Client 初始化成功，已兼容 Oracle 11g！")
    except Exception as e:
        print(f"   ⚠️ Oracle Client 初始化报错: {e}")

    print("3. 正在连接 Oracle 数据库...")
    conn = oracledb.connect(**DB_CONFIG)
    cursor = conn.cursor()
    print("   ✅ 数据库连接成功！")

    print("4. 开始按二级分类建文件夹并导出 Excel...")

    for level2_name, l3_dict in schema.items():
        # 【二级分类创建独立文件夹】
        clean_l2_folder_name = sanitize_filename(level2_name)
        l2_folder_path = os.path.join(BASE_OUTPUT_DIR, clean_l2_folder_name)

        if not os.path.exists(l2_folder_path):
            os.makedirs(l2_folder_path)
            print(f"\n📁 [创建二级分类文件夹] {clean_l2_folder_name}")
        else:
            print(f"\n📁 [二级分类文件夹已存在] {clean_l2_folder_name}")

        # 遍历三级小类生成 Excel
        for level3_name, table_list in l3_dict.items():
            if table_list:
                print(f"  └── 正在处理三级小类: {level3_name}")
                generate_excel_for_l3(
                    level2_name, level3_name, table_list, cursor, l2_folder_path
                )

    cursor.close()
    conn.close()
    print("\n🎉 处理全部完成！所有文件与文件夹已生成在目标路径下。")


if __name__ == "__main__":
    main()